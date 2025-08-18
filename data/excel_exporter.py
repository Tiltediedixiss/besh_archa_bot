import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os
from typing import Dict, List, Any
from .progress import progress_manager

class ExcelExporter:
    """Класс для экспорта данных о прогрессе пользователей в Excel"""
    
    def __init__(self, output_dir: str = "reports"):
        self.output_dir = output_dir
        self.ensure_output_dir()
    
    def ensure_output_dir(self):
        """Создает директорию для отчетов, если она не существует"""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
    
    def export_user_progress(self, filename: str = None) -> str:
        """Экспортирует прогресс всех пользователей в Excel файл"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"user_progress_report_{timestamp}.xlsx"
        
        filepath = os.path.join(self.output_dir, filename)
        
        # Получаем данные о прогрессе всех пользователей
        all_progress = self._get_all_user_progress()
        
        # Создаем Excel файл
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Основная таблица прогресса
            self._create_main_progress_sheet(writer, all_progress)
            
            # Детальная таблица по модулям
            self._create_modules_detail_sheet(writer, all_progress)
            
            # Статистика по дням
            self._create_days_statistics_sheet(writer, all_progress)
            
            # Сводка по завершению
            self._create_completion_summary_sheet(writer, all_progress)
        
        return filepath
    
    def _get_all_user_progress(self) -> List[Dict[str, Any]]:
        """Получает прогресс всех пользователей"""
        return progress_manager.get_all_users_progress()
    
    def _create_main_progress_sheet(self, writer: pd.ExcelWriter, data: List[Dict[str, Any]]):
        """Создает основной лист с прогрессом пользователей"""
        if not data:
            # Пустой отчет без демонстрационных данных
            df = pd.DataFrame(columns=[
                'User ID', 'Username', 'First Name', 'Last Name', 'Current Day',
                'Day 1 Status', 'Day 2 Status', 'Day 3 Status', 'Day 4 Status',
                'Total Modules Completed', 'Last Activity', 'Is Admin'
            ])
        else:
            df = pd.DataFrame(data)
        
        df.to_excel(writer, sheet_name='Основной прогресс', index=False)
        
        # Применяем стили
        self._apply_main_sheet_styling(writer.sheets['Основной прогресс'])
    
    def _create_modules_detail_sheet(self, writer: pd.ExcelWriter, data: List[Dict[str, Any]]):
        """Создает детальный лист по модулям"""
        # Получаем детальные данные по модулям
        modules_data = progress_manager.get_modules_detail_data()
        
        if not modules_data:
            df = pd.DataFrame(columns=[
                'User ID', 'Username', 'Day', 'Module', 'Status',
                'Completion Date', 'Time Spent (min)'
            ])
        else:
            df = pd.DataFrame(modules_data)
        
        df.to_excel(writer, sheet_name='Детали по модулям', index=False)
        
        # Применяем стили
        self._apply_modules_sheet_styling(writer.sheets['Детали по модулям'])
    
    def _create_days_statistics_sheet(self, writer: pd.ExcelWriter, data: List[Dict[str, Any]]):
        """Создает лист со статистикой по дням"""
        # Получаем статистику по дням
        days_stats = progress_manager.get_days_statistics()
        
        if not days_stats:
            df = pd.DataFrame(columns=[
                'Day', 'Total Users', 'Completed', 'In Progress', 'Locked',
                'Completion Rate (%)', 'Average Time (min)'
            ])
        else:
            df = pd.DataFrame(days_stats)
        
        df.to_excel(writer, sheet_name='Статистика по дням', index=False)
        
        # Применяем стили
        self._apply_statistics_sheet_styling(writer.sheets['Статистика по дням'])
    
    def _create_completion_summary_sheet(self, writer: pd.ExcelWriter, data: List[Dict[str, Any]]):
        """Создает лист со сводкой по завершению"""
        # Получаем сводку по завершению
        completion_summary = progress_manager.get_completion_summary()
        
        if not completion_summary:
            df = pd.DataFrame(columns=['Metric', 'Value', 'Notes'])
        else:
            df = pd.DataFrame(completion_summary)
        
        df.to_excel(writer, sheet_name='Сводка по завершению', index=False)
        
        # Применяем стили
        self._apply_summary_sheet_styling(writer.sheets['Сводка по завершению'])
    
    def _apply_main_sheet_styling(self, worksheet):
        """Применяет стили к основному листу"""
        # Заголовки
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Автоматическая ширина колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _apply_modules_sheet_styling(self, worksheet):
        """Применяет стили к листу модулей"""
        # Заголовки
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Автоматическая ширина колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _apply_statistics_sheet_styling(self, worksheet):
        """Применяет стили к листу статистики"""
        # Заголовки
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Автоматическая ширина колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _apply_summary_sheet_styling(self, worksheet):
        """Применяет стили к листу сводки"""
        # Заголовки
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Автоматическая ширина колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

# Создаем глобальный экземпляр экспортера
excel_exporter = ExcelExporter() 